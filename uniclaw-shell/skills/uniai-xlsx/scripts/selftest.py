"""自检：用 helper 造样张，跑规则检查，验证关键不变量与文档一致性。

    python scripts/selftest.py            # 临时目录，跑完即删
    python scripts/selftest.py --keep out # 保留样张

改动 xlsx_helpers.py / check_xlsx.py / recalc.py / SKILL.md 之后跑这个——
它是"文档说的"和"代码做的"开始分叉时的报警器。不依赖 LibreOffice。
"""

from __future__ import annotations

import argparse
import sys
import tempfile
from os.path import abspath, dirname, exists, join

sys.path.insert(0, dirname(abspath(__file__)))

from openpyxl import Workbook, load_workbook                # noqa: E402

import check_xlsx as ck                                     # noqa: E402
from office.soffice import find_soffice, soffice_available, user_profile_dir  # noqa: E402
from xlsx_helpers import (FMT, THEMES, add_title, add_total_row,   # noqa: E402
                          autofit_columns, style_table, write_table)

FAILURES: list[str] = []
DATA = [["核心业务", 12480], ["新兴业务", 5320], ["海外业务", 2910],
        ["服务与运维", 1540], ["硬件销售", 520], ["其他", 250],
        ["培训服务", 180], ["其他收入", 120]]


def check(cond, msg):
    print(("  ok   " if cond else "  FAIL ") + msg)
    if not cond:
        FAILURES.append(msg)


def build(path, theme="navy"):
    wb = Workbook()
    ws = wb.active
    ws.title = "收入"
    add_title(ws, "2026 年上半年收入分析", "数据来源：公司财报", span=3, theme=theme)
    rows = [[name, val, f"=B{4 + i}/$B${4 + len(DATA)}"] for i, (name, val) in enumerate(DATA)]
    h, f, l = write_table(ws, ["业务线", "收入（万元）", "占比"], rows, start_row=3)
    tr = add_total_row(ws, f, l, cols="B")
    ws[f"C{tr}"] = f"=SUM(C{f}:C{l})"
    style_table(ws, header_row=h, first_row=f, last_row=l, total_row=tr,
                money_cols="B", percent_cols="C", theme=theme)
    autofit_columns(ws)
    wb.save(path)
    return path, h, f, l, tr


def rules_pass(path):
    rep = ck.Report()
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        ck.check_sheet(ws, rep)
    return ([i for i in rep.items if i["level"] == "ERROR"],
            [i for i in rep.items if i["level"] == "WARN"])


def invariants(path, h, f, l, tr):
    ws = load_workbook(path)["收入"]

    check(str(ws[f"B{tr}"].value).startswith("=SUM("), "合计行是公式，不是算好的数字")
    check(ws[f"B{tr}"].alignment.horizontal == "right",
          "合计行（公式单元格）也被右对齐了——按列判定，不按 isinstance")
    check(ws[f"B{f}"].number_format == FMT["money"], "金额列设了千分位格式")
    check(ws[f"C{f}"].number_format == FMT["percent"], "占比列设了百分比格式")
    check(ws.freeze_panes == f"A{h + 1}", "冻结窗格落在表头下一行")

    hdr = ws.cell(row=h, column=1)
    check(ck.rgb_of(hdr.fill.fgColor) == THEMES["navy"]["accent"], "表头是强调色底")
    check(hdr.font.bold and ck.rgb_of(hdr.font.color) == "FFFFFF", "表头白字加粗")
    check(all("微软雅黑" in (ws.cell(row=r, column=1).font.name or "")
              for r in range(f, l + 1)), "中文用中文字体，不是 Arial")

    widths = {k: v.width for k, v in ws.column_dimensions.items() if v.width}
    check(widths.get("A", 0) < 20,
          f"标题合并单元格没有把 A 列撑宽（实得 {widths.get('A', 0):.1f}）")
    check(all(w >= 8 for w in widths.values()), "所有列都设了不小于 8 的列宽")
    check(not [m for m in ws.merged_cells.ranges if m.min_row >= f],
          "数据区没有合并单元格")


def recalc_degradation():
    """没有 LibreOffice 时必须给出结构化兜底，而不是抛 traceback。"""
    import recalc
    if soffice_available():
        check(True, f"本机有 LibreOffice（{find_soffice()}），跳过降级路径测试")
        return
    payload = recalc.recalc(join(dirname(abspath(__file__)), "..", "SKILL.md"))
    check(payload.get("status") in ("no_soffice", "error"),
          f"缺 LibreOffice 时返回结构化结果而非崩溃（status={payload.get('status')}）")
    check("fallback" in payload or "error" in payload, "给出了兜底指引")
    check("/user" in user_profile_dir() or "user" in user_profile_dir(),
          f"用户配置目录按平台解析：{user_profile_dir()}")


def skill_doc_consistency():
    root = dirname(dirname(abspath(__file__)))
    skill = open(join(root, "SKILL.md"), encoding="utf-8").read()
    import xlsx_helpers as xh
    for name in ("write_table", "add_total_row", "style_table", "autofit_columns",
                 "add_title", "mark_input"):
        if name in skill:
            check(hasattr(xh, name), f"SKILL.md 提到的 {name}() 存在")
    for f in ("scripts/xlsx_helpers.py", "scripts/check_xlsx.py", "scripts/recalc.py"):
        if f in skill:
            check(exists(join(root, f)), f"SKILL.md 引用的 {f} 存在")
    for theme in THEMES:
        check(THEMES[theme]["accent"] in skill, f"主题 {theme} 的色值在 SKILL.md 里有")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--keep", default=None)
    args = ap.parse_args()
    import os
    tmp = args.keep or tempfile.mkdtemp(prefix="xlsx_selftest_")
    os.makedirs(tmp, exist_ok=True)

    print("SKILL.md ↔ 代码一致性")
    skill_doc_consistency()

    print("\nrecalc 降级路径")
    recalc_degradation()

    for theme in THEMES:
        path = join(tmp, f"sample_{theme}.xlsx")
        print(f"\ntheme={theme} → {path}")
        _, h, f, l, tr = build(path, theme)
        errors, warns = rules_pass(path)
        check(not errors, "check_xlsx 无 ERROR" +
              ("" if not errors else "：" + "；".join(e["code"] for e in errors)))
        if warns:
            print("       (warn: " + "、".join(w["code"] for w in warns) + ")")
        if theme == "navy":
            invariants(path, h, f, l, tr)

    print()
    if FAILURES:
        print(f"✗ {len(FAILURES)} 项未通过：")
        for x in FAILURES:
            print("   - " + x)
        sys.exit(1)
    print(f"✓ 全部通过。样张在 {tmp}" + ("" if args.keep else "（临时目录）"))


if __name__ == "__main__":
    main()
