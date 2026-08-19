"""openpyxl 报表助手：把 SKILL.md 的硬规则变成可调用的函数。

    import sys; sys.path.insert(0, '<skill_dir>/scripts')
    from xlsx_helpers import *

    wb = Workbook(); ws = wb.active
    write_table(ws, header=['业务线','收入（万元）','占比'], rows=data, start_row=3)
    add_total_row(ws, first_row=4, last_row=9, cols='BC', label_col='A')
    style_table(ws, header_row=3, first_row=4, last_row=10, theme='navy',
                money_cols='B', percent_cols='C')
    autofit_columns(ws)

为什么要用这些函数：SKILL.md 里那段"数字右对齐"的片段有个隐蔽 bug——
`isinstance(cell.value, (int, float))` 对公式单元格是**假**（公式的值是字符串 '=SUM(...)'），
所以合计行永远不会被右对齐，而技能自己写着"数字列左对齐 = 不合格"。
这里的 `style_table` 按**列**判定，不按单个单元格的 Python 类型判定。
"""

from __future__ import annotations

import re

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string

__all__ = [
    "THEMES", "INK", "MUTED", "LINE", "ZEBRA", "FMT",
    "write_table", "add_total_row", "style_table", "autofit_columns",
    "style_header", "set_number_format", "add_title", "mark_input",
    "mark_formula", "mark_link", "add_source_note", "col_letters",
    "Font", "PatternFill", "Alignment", "Border", "Side", "Comment",
]

# 与 uniai-docx 同一套主题色，两个技能出的文档放在一起才像一家人
THEMES = {
    "navy":   {"accent": "1F4E79", "tint": "EAF1F8", "name": "藏青·通用首选"},
    "green":  {"accent": "1E5C40", "tint": "E8F2EC", "name": "墨绿·金融审计"},
    "teal":   {"accent": "0F766E", "tint": "E5F2F0", "name": "青碧·科技数据"},
    "ochre":  {"accent": "9A3412", "tint": "FAEEE7", "name": "赭石·消费文教"},
    "gray":   {"accent": "404040", "tint": "F2F2F2", "name": "深灰·极简中性"},
}
INK = "1F2937"       # 正文墨色，别用纯黑
MUTED = "595959"     # 副题、注释
LINE = "D9D9D9"      # 细边框
ZEBRA = "F7F8FA"     # 隔行底色

FMT = {
    "money": "#,##0",
    "money2": "#,##0.00",
    "money_neg": "#,##0;(#,##0);-",     # 负数括号、零显示为 -
    "percent": "0.0%",
    "percent0": "0%",
    "int": "#,##0",
    "date": "yyyy-mm-dd",
    "text": "@",
}

CJK = re.compile(r"[㐀-鿿豈-﫿　-〿＀-￯]")
DEFAULT_FONT = "微软雅黑"


def col_letters(cols):
    """'BC' / ['B','C'] / [2,3] 统一成 ['B','C']。"""
    if isinstance(cols, str):
        return list(cols)
    return [c if isinstance(c, str) else get_column_letter(c) for c in cols]


def _font(size=10.5, bold=False, color=INK, name=DEFAULT_FONT):
    return Font(name=name, size=size, bold=bold, color=color)


# --- 写入 -------------------------------------------------------------------
def add_title(ws, title, subtitle=None, row=1, span=6, theme="navy"):
    """报表标题区：标题 14-16pt 加粗 + 副题灰字，与数据区空一行。"""
    t = THEMES[theme] if isinstance(theme, str) else theme
    ws.cell(row=row, column=1, value=title).font = _font(15, True, t["accent"])
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(2, span))
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = _font(9, False, MUTED)
        ws.merge_cells(start_row=row + 1, start_column=1,
                       end_row=row + 1, end_column=max(2, span))
        ws.row_dimensions[row + 1].height = 16
    return ws


def write_table(ws, header, rows, start_row=1, start_col=1):
    """写表头 + 数据。返回 (header_row, first_data_row, last_data_row)。"""
    for j, name in enumerate(header):
        ws.cell(row=start_row, column=start_col + j, value=name)
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            ws.cell(row=start_row + 1 + i, column=start_col + j, value=v)
    return start_row, start_row + 1, start_row + len(rows)


def add_total_row(ws, first_row, last_row, cols, label_col="A", label="合计", row=None):
    """合计行——**写公式，不写算好的数**。返回合计行行号。"""
    row = row or last_row + 1
    ws.cell(row=row, column=column_index_from_string(label_col), value=label)
    for c in col_letters(cols):
        ws[f"{c}{row}"] = f"=SUM({c}{first_row}:{c}{last_row})"
    return row


def set_number_format(ws, cols, fmt, first_row, last_row):
    """整列设数字格式。fmt 用 FMT 里的键或原始格式串。"""
    fmt = FMT.get(fmt, fmt)
    for c in col_letters(cols):
        for r in range(first_row, last_row + 1):
            ws[f"{c}{r}"].number_format = fmt
    return ws


# --- 样式 -------------------------------------------------------------------
def style_header(ws, row, first_col=1, last_col=None, theme="navy", height=22):
    t = THEMES[theme] if isinstance(theme, str) else theme
    last_col = last_col or ws.max_column
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _font(11, True, "FFFFFF")
        cell.fill = PatternFill("solid", fgColor=t["accent"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height
    return ws


def style_table(ws, header_row, first_row, last_row, theme="navy",
                first_col=1, last_col=None, money_cols=(), percent_cols=(),
                text_cols=(), date_cols=(), total_row=None, zebra=None,
                freeze=True, font_size=10.5, row_height=18):
    """一次把表做完：表头着色、数字格式、对齐、细边框、斑马纹、冻结、合计行加粗。

    **按列判定对齐**——公式单元格的值是字符串，按单元格类型判会漏掉整个合计行。
    """
    t = THEMES[theme] if isinstance(theme, str) else theme
    last_col = last_col or ws.max_column
    if zebra is None:
        zebra = (last_row - first_row + 1) > 7

    style_header(ws, header_row, first_col, last_col, theme)
    set_number_format(ws, money_cols, "money", first_row, total_row or last_row)
    set_number_format(ws, percent_cols, "percent", first_row, total_row or last_row)
    set_number_format(ws, text_cols, "text", first_row, total_row or last_row)
    set_number_format(ws, date_cols, "date", first_row, total_row or last_row)

    right = set(col_letters(money_cols)) | set(col_letters(percent_cols))
    center = set(col_letters(date_cols))
    thin = Side(style="thin", color=LINE)
    end = total_row or last_row

    for r in range(first_row, end + 1):
        ws.row_dimensions[r].height = row_height
        is_total = total_row is not None and r == total_row
        for c in range(first_col, last_col + 1):
            letter = get_column_letter(c)
            cell = ws.cell(row=r, column=c)
            cell.font = _font(font_size, bold=is_total)
            cell.border = Border(bottom=thin)
            if letter in right:
                align = "right"
            elif letter in center:
                align = "center"
            else:
                align = "left"
            cell.alignment = Alignment(horizontal=align, vertical="center",
                                       wrap_text=False)
            if is_total:
                cell.fill = PatternFill("solid", fgColor=t["tint"])
            elif zebra and (r - first_row) % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
    if freeze:
        ws.freeze_panes = f"A{header_row + 1}"
    return ws


def autofit_columns(ws, min_width=8, max_width=48, padding=2.5):
    """按内容估列宽——中文按 2 倍宽算，不裁字也不留大片空白。"""
    # 合并单元格（标题区）不能参与列宽估算，否则一个长标题会把 A 列撑到几十字符宽
    merged = {coord for rng in ws.merged_cells.ranges for coord in rng.coord.split(":")}
    merged |= {c.coordinate for rng in ws.merged_cells.ranges for row in rng.cells
               for c in [ws.cell(row=row[0], column=row[1])]}

    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.coordinate in merged:
                continue
            text = str(cell.value)
            if text.startswith("="):
                text = "0" * 12          # 公式按结果宽度估，别按公式串长度
            w = sum(2 if CJK.search(ch) else 1 for ch in text)
            letter = get_column_letter(cell.column)
            widths[letter] = max(widths.get(letter, 0), w)
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = max(min_width, min(max_width, w + padding))
    return ws


# --- 财务模型配色（行业惯例） -----------------------------------------------
def mark_input(cell, note=None):
    """蓝字 = 硬编码输入/假设。附来源批注是硬要求。"""
    cell.font = Font(name=DEFAULT_FONT, size=10.5, color="0000FF")
    if note:
        add_source_note(cell, note)
    return cell


def mark_formula(cell):
    cell.font = Font(name=DEFAULT_FONT, size=10.5, color="000000")
    return cell


def mark_link(cell, external=False):
    """绿字 = 同簿跨表引用；红字 = 外部文件引用。"""
    cell.font = Font(name=DEFAULT_FONT, size=10.5,
                     color="FF0000" if external else "008000")
    return cell


def add_source_note(cell, text, author="skill"):
    cell.comment = Comment(text, author)
    return cell
