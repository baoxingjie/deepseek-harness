"""重算工作簿里的公式并扫描错误（依赖 LibreOffice）。

    python scripts/recalc.py output.xlsx [timeout_seconds]

openpyxl 写进去的公式只是字符串、没有计算结果，所以用了公式就必须跑这一步。
输出 JSON：

    {"status":"success","total_errors":0,"total_formulas":42,"recalculated":true}
    {"status":"errors_found","error_summary":{"#REF!":{"count":2,"locations":[...]}}}
    {"status":"no_soffice","fallback":"..."}          ← 没装 LibreOffice，走兜底
    {"status":"not_recalculated","hint":"..."}        ← soffice 跑了但公式没被算出来

`status` 不是 success 时**不要假装重算成功了**——按 SKILL.md 的兜底路径告知用户。
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from office.soffice import (
    SofficeNotFound,
    find_soffice,
    get_soffice_env,
    user_profile_dir,
)

from openpyxl import load_workbook

MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]

FALLBACK = (
    "本机没有 LibreOffice，无法重算。两条路，二选一并**明确告诉用户**："
    "① 交付文件，说明「用 Excel/WPS 打开会自动重算」，并确保公式本身正确；"
    "② 如果对方不会用 Excel 打开、结果值必须内嵌，就在 Python 侧算好写成静态值，"
    "并注明这是静态快照。不要假装重算成功了。"
)


def has_gtimeout():
    try:
        subprocess.run(["gtimeout", "--version"], capture_output=True, timeout=1, check=False)
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def macro_path():
    return os.path.join(user_profile_dir(), "basic", "Standard", MACRO_FILENAME)


def setup_libreoffice_macro(soffice):
    """把重算宏装进用户配置目录（三平台路径都对）。"""
    macro_file = macro_path()
    macro_dir = os.path.dirname(macro_file)

    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True

    if not os.path.exists(macro_dir):
        # 先让 LibreOffice 自己把用户配置目录建出来
        subprocess.run(
            [soffice, "--headless", "--terminate_after_init"],
            capture_output=True, timeout=60, check=False, env=get_soffice_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except OSError:
        return False


def scan_errors(filename):
    """扫描缓存值里的 Excel 错误，并判断重算到底有没有发生。"""
    wb = load_workbook(filename, data_only=True)
    details = {err: [] for err in EXCEL_ERRORS}
    total = 0
    cached_values = 0
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cached_values += 1
                if isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            details[err].append(f"{name}!{cell.coordinate}")
                            total += 1
                            break
    wb.close()

    wb_f = load_workbook(filename, data_only=False)
    formulas = 0
    formula_cells = []
    for name in wb_f.sheetnames:
        for row in wb_f[name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                    formula_cells.append((name, cell.coordinate))
    wb_f.close()

    # 公式格在 data_only 视图里全是 None → 根本没被算过（soffice 静默失败时就是这样）
    recalculated = True
    if formulas:
        wb_c = load_workbook(filename, data_only=True)
        got = sum(1 for name, coord in formula_cells if wb_c[name][coord].value is not None)
        wb_c.close()
        recalculated = got > 0
    return details, total, formulas, recalculated, cached_values


def result_payload(filename):
    details, total, formulas, recalculated, _ = scan_errors(filename)
    payload = {
        "status": "success" if total == 0 else "errors_found",
        "total_errors": total,
        "total_formulas": formulas,
        "recalculated": recalculated,
        "error_summary": {},
    }
    for err, locs in details.items():
        if locs:
            payload["error_summary"][err] = {"count": len(locs), "locations": locs[:20]}
    if formulas and not recalculated:
        payload["status"] = "not_recalculated"
        payload["hint"] = (
            f"{formulas} 个公式单元格都没有缓存值——LibreOffice 没有真正重算（可能启动失败或超时）。"
            "别把这个结果当成「零错误」。" + FALLBACK
        )
    return payload


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"status": "error", "error": f"文件不存在：{filename}"}

    try:
        soffice = find_soffice()
    except SofficeNotFound as exc:
        payload = {"status": "no_soffice", "error": str(exc), "fallback": FALLBACK}
        try:    # 即便不能重算，公式数量和"上一次的缓存值里有没有错误"仍然值得报
            details, total, formulas, recalculated, _ = scan_errors(filename)
            payload.update(total_formulas=formulas, cached_errors=total, recalculated=False)
        except Exception:
            pass
        return payload

    if not setup_libreoffice_macro(soffice):
        return {"status": "error",
                "error": f"无法写入重算宏：{macro_path()}（检查目录权限）"}

    cmd = [
        soffice, "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        str(Path(filename).absolute()),
    ]
    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd

    try:
        proc = subprocess.run(cmd, capture_output=True, text=True,
                              env=get_soffice_env(), timeout=timeout + 30)
    except subprocess.TimeoutExpired:
        return {"status": "error", "error": f"LibreOffice 超时（>{timeout}s）", "fallback": FALLBACK}

    # returncode 124 = timeout 命令自己的超时码，文件通常已经存过了，继续往下检
    if proc.returncode not in (0, 124):
        return {"status": "error",
                "error": (proc.stderr or "").strip() or f"LibreOffice 退出码 {proc.returncode}",
                "fallback": FALLBACK}

    try:
        return result_payload(filename)
    except Exception as exc:                     # noqa: BLE001
        return {"status": "error", "error": f"读取重算结果失败：{exc}"}


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    payload = recalc(filename, timeout)
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    sys.exit(0 if payload.get("status") == "success" else 1)


if __name__ == "__main__":
    main()
