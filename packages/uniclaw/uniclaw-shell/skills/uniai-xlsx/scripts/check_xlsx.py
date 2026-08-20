"""XLSX 交付前自检：把 SKILL.md 的硬规则跑成一道闸门。

    python scripts/check_xlsx.py out/report.xlsx
    python scripts/check_xlsx.py out/report.xlsx --json --strict

有 ERROR 就是返工（退出码 1）。它查的是"打开看一眼像那么回事、细看全是问题"的那些：
派生值被硬编码成数字、公式错误、裸表头、数字列左对齐、没设数字格式、没冻结、
中文用西文字体、列宽裁字、数据区滥用合并单元格、强调色不止一个。

它**不能**替代 `recalc.py`：公式算不算得出来、有没有 #REF!，要靠 LibreOffice 真算一遍。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CJK = re.compile(r"[㐀-鿿豈-﫿　-〿＀-￯]")
ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")
# 这些表头/行标一出现，该列或该行的值就应该是公式而不是常数
DERIVED_WORDS = ("合计", "总计", "小计", "累计", "平均", "占比", "比例", "增速",
                 "同比", "环比", "增长率", "毛利", "利润率", "差额", "total", "sum",
                 "average", "growth", "margin", "%")
NEUTRAL = {"000000", "FFFFFF", "1F2937", "595959", "D9D9D9", "F7F8FA", "F2F2F2",
           "E7E6E6", "808080", "404040", "BFBFBF", "auto", None}
WESTERN_ONLY = {"Arial", "Calibri", "Times New Roman", "Helvetica", "Verdana", "Tahoma"}


class Report:
    def __init__(self):
        self.items = []

    def add(self, level, code, msg, where=""):
        self.items.append({"level": level, "code": code, "msg": msg, "where": where})

    error = lambda self, c, m, w="": self.add("ERROR", c, m, w)
    warn = lambda self, c, m, w="": self.add("WARN", c, m, w)
    info = lambda self, c, m, w="": self.add("INFO", c, m, w)


def rgb_of(color):
    """openpyxl 的颜色可能是 str、RGB 对象或主题色，统一成 6 位十六进制或 None。"""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb is None or not isinstance(rgb, str):   # RGB 对象 / 主题色 / indexed
        return None
    return rgb[-6:].upper()


def is_derived_label(text):
    t = str(text or "").strip().lower()
    return any(w.lower() in t for w in DERIVED_WORDS)


def sheet_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                yield cell


def find_header_row(ws, limit=8):
    """表头 = 前几行里非空单元格最多、且基本都是文本的那一行。"""
    best, best_score = None, 0
    for r in range(1, min(ws.max_row, limit) + 1):
        cells = [c for c in ws[r] if c.value is not None]
        if len(cells) < 2:
            continue
        texts = sum(1 for c in cells if isinstance(c.value, str) and not str(c.value).startswith("="))
        if texts == len(cells) and len(cells) > best_score:
            best, best_score = r, len(cells)
    return best


def check_sheet(ws, rep):
    where = ws.title
    header_row = find_header_row(ws)
    if header_row is None:
        rep.info("NO_TABLE", "没识别出表头行，跳过表格类检查（可能是纯说明页）。", where)
        return

    first_data = header_row + 1
    ncols = ws.max_column

    # --- 1. 派生值被硬编码 ------------------------------------------------
    hardcoded = []
    derived_cols = {c for c in range(1, ncols + 1)
                    if is_derived_label(ws.cell(row=header_row, column=c).value)}
    for r in range(first_data, ws.max_row + 1):
        row_label = str(ws.cell(row=r, column=1).value or "")
        row_is_derived = is_derived_label(row_label)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell.value, (int, float)) or isinstance(cell.value, bool):
                continue
            if row_is_derived or c in derived_cols:
                hardcoded.append(cell.coordinate)
    if hardcoded:
        rep.error("HARDCODED_DERIVED",
                  f"{len(hardcoded)} 个派生值是写死的数字，不是公式："
                  f"{'、'.join(hardcoded[:8])}{'…' if len(hardcoded) > 8 else ''}。"
                  "合计/占比/增速这类值必须写成 Excel 公式，源数据一改就得跟着变。", where)

    # --- 2. 缓存里的公式错误 ---------------------------------------------
    errs = Counter()
    for cell in sheet_cells(ws):
        if isinstance(cell.value, str):
            for e in ERRORS:
                if e in cell.value:
                    errs[e] += 1
                    break
    if errs:
        rep.error("FORMULA_ERROR",
                  "存在公式错误：" + "、".join(f"{k}×{v}" for k, v in errs.items())
                  + "。交付前必须清零。", where)

    # --- 3. 表头 ----------------------------------------------------------
    hdr = [ws.cell(row=header_row, column=c) for c in range(1, ncols + 1)]
    filled = any(rgb_of(c.fill.fgColor) not in (None, "000000", "FFFFFF")
                 for c in hdr if c.fill is not None)
    bold = any(c.font is not None and c.font.bold for c in hdr)
    if not filled and not bold:
        rep.error("BARE_HEADER", f"第 {header_row} 行是裸表头：既没底色也没加粗。", where)
    elif not filled:
        rep.warn("BARE_HEADER", "表头只加粗没底色，层级偏弱。", where)

    if not ws.freeze_panes:
        rep.warn("NO_FREEZE", f"没有冻结窗格，滚动就看不到表头了（freeze_panes='A{first_data}'）。",
                 where)

    # --- 4. 数字列：格式与对齐 -------------------------------------------
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        vals, fmts, aligns = [], set(), set()
        for r in range(first_data, ws.max_row + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            numeric = isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
            formula = isinstance(cell.value, str) and cell.value.startswith("=")
            if numeric or formula:
                vals.append(cell)
                fmts.add(cell.number_format)
                aligns.add(cell.alignment.horizontal if cell.alignment else None)
        if len(vals) < 2:
            continue
        if aligns and all(a in (None, "left", "general") for a in aligns):
            rep.error("NUM_ALIGN", f"{letter} 列是数字列却左对齐（公式单元格最容易漏，"
                                   "别按 isinstance 判类型，按列判）。", where)
        if fmts <= {"General"}:
            rep.warn("NO_NUMBER_FORMAT",
                     f"{letter} 列没设数字格式，金额不带千分位、百分比显示成小数。", where)

    # --- 5. 字体与颜色 ----------------------------------------------------
    accent, cjk_western = set(), []
    for cell in sheet_cells(ws):
        f = cell.font
        if f is not None:
            fc = rgb_of(f.color)
            # 蓝/绿/红是财务模型的语义色（输入/跨表/外部引用），不算"强调色"
            if fc and fc not in NEUTRAL and fc not in ("0000FF", "008000", "FF0000"):
                accent.add(fc)
            if CJK.search(str(cell.value)) and f.name in WESTERN_ONLY:
                cjk_western.append(cell.coordinate)
        if cell.fill is not None:
            bg = rgb_of(cell.fill.fgColor)
            if bg and bg not in NEUTRAL and bg != "000000":
                accent.add(bg)
    if cjk_western:
        rep.error("CJK_FONT",
                  f"{len(cjk_western)} 个含中文的单元格用了西文字体（{cjk_western[:5]}），"
                  "中文会串码或字形难看，改用微软雅黑/宋体。", where)
    if len(accent) > 2:
        rep.warn("MULTI_ACCENT",
                 f"出现 {len(accent)} 种非中性色：{sorted(accent)[:6]}。"
                 "全簿只锁一个强调色 + 配套浅底。", where)

    # --- 6. 列宽与合并单元格 ---------------------------------------------
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        need = 0
        for r in range(header_row, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or (isinstance(v, str) and v.startswith("=")):
                continue
            need = max(need, sum(2 if CJK.search(ch) else 1 for ch in str(v)))
        if not need:                 # 空列（多半是标题合并跨过去的），不参与列宽检查
            continue
        dim = ws.column_dimensions.get(letter)
        width = dim.width if dim is not None and dim.width else None
        if width is None:
            rep.warn("NO_COL_WIDTH", f"{letter} 列没设列宽，默认 8.43 常常裁字。", where)
        elif need and width < need * 0.75:
            rep.warn("COL_TOO_NARROW",
                     f"{letter} 列宽 {width:.0f}，内容最宽约 {need} 字符，会被裁掉。", where)

    data_merges = [str(m) for m in ws.merged_cells.ranges if m.min_row >= first_data]
    if data_merges:
        rep.error("MERGED_IN_DATA",
                  f"数据区有合并单元格（{data_merges[:5]}），会毁掉排序、筛选和公式引用。"
                  "合并只用于标题区。", where)


def main():
    ap = argparse.ArgumentParser(description="XLSX 交付前自检")
    ap.add_argument("path")
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--strict", action="store_true", help="WARN 也算失败")
    args = ap.parse_args()

    rep = Report()
    wb = load_workbook(args.path, data_only=False)
    for ws in wb.worksheets:
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            continue
        check_sheet(ws, rep)

    # 缓存值视角：公式有没有被算过
    wbc = load_workbook(args.path, data_only=True)
    formulas = sum(1 for ws in wb.worksheets for c in sheet_cells(ws)
                   if isinstance(c.value, str) and c.value.startswith("="))
    if formulas:
        cached = sum(1 for ws in wb.worksheets for c in sheet_cells(ws)
                     if isinstance(c.value, str) and c.value.startswith("=")
                     and wbc[ws.title][c.coordinate].value is not None)
        if cached == 0:
            rep.warn("NOT_RECALCULATED",
                     f"{formulas} 个公式都没有缓存值——还没跑过 recalc.py。"
                     "交付前要么重算，要么明确告诉用户「用 Excel 打开会自动计算」。")

    errors = [i for i in rep.items if i["level"] == "ERROR"]
    warns = [i for i in rep.items if i["level"] == "WARN"]

    if args.json:
        print(json.dumps({"path": args.path, "errors": len(errors),
                          "warnings": len(warns), "items": rep.items},
                         ensure_ascii=False, indent=2))
    else:
        icon = {"ERROR": "✗", "WARN": "!", "INFO": "·"}
        for lvl in ("ERROR", "WARN", "INFO"):
            for it in [i for i in rep.items if i["level"] == lvl]:
                loc = f" [{it['where']}]" if it["where"] else ""
                print(f"{icon[lvl]} {lvl:<5} {it['code']}{loc}: {it['msg']}")
        print(f"\n{len(errors)} error(s), {len(warns)} warning(s) — "
              f"{len(wb.worksheets)} 个 sheet，{formulas} 个公式")
        if not rep.items:
            print("干净。公式是否算得出来仍要靠 recalc.py 验。")

    sys.exit(1 if errors or (args.strict and warns) else 0)


if __name__ == "__main__":
    main()
